import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

# Task 1: Calculate how many Products we have per Supplier
products_per_supplier = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value

    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier[supplier_name] # or you can use products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        products_per_supplier[supplier_name] = 1

print(f"Task 1: {products_per_supplier}")

# Task 2: List each company with respective total Inventory value
total_value_per_supplier = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value

    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else: 
        total_value_per_supplier[supplier_name] = inventory * price

print(f"Task 2: {total_value_per_supplier}")

#Task 3: List Products with Inventory less than 10
products_under_10_inv = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value

    if inventory < 10: 
        products_under_10_inv[int(product_num)] = int(inventory)

print(f"Task 3: {products_under_10_inv}")

#Task 4: Add a fifth column to the spreadsheetnfor total inventory price

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    inventory_price.value = inventory * price

inv_file.save("inventory_with_total_value.xlsx")








